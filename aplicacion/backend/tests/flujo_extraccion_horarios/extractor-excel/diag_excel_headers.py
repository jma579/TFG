# diag_excel_headers.py
from pathlib import Path
from openpyxl import load_workbook
import unicodedata, re, sys

DAYS = ["LUNES","MARTES","MIÉRCOLES","JUEVES","VIERNES"]

def norm(s):
    s = (s or "").strip().upper()
    s = " ".join(s.split())
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s

def main(xlsx):
    wb = load_workbook(xlsx, data_only=True)
    for ws in wb.worksheets:
        print(f"\n== Hoja: {ws.title} ==")
        found = False
        for r in range(1, min(ws.max_row, 200)+1):
            row = [norm(c.value) for c in ws[r]]
            cols = []
            for d in DAYS:
                cidx = [i for i,v in enumerate(row, start=1) if v == d]
                cols.append((d, cidx))
            # Imprime filas con todos los días presentes al menos una vez
            if all(len(cidx)>0 for _, cidx in cols):
                found = True
                print(f"Fila {r}: ", {d: c for d, c in cols})
        if not found:
            print("  (no se detectaron cabeceras en las primeras 200 filas)")

if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv)>1 else r"D:\TFG\Horarios\Grado\HORARIOS 2025-26.xlsx"
    main(xlsx)
