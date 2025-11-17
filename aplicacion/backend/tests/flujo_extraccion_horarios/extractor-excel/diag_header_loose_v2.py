# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook
import unicodedata, re, sys

DAYS = ["LUNES","MARTES","MIÉRCOLES","JUEVES","VIERNES"]

def norm(s: str) -> str:
    s = (s or "").strip().upper()
    s = " ".join(s.split())
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s

def day_token_match(txt: str, canonical: str) -> bool:
    s = norm(txt); d = norm(canonical)
    if not s or not d: return False
    if s == d: return True
    return re.match(rf"^{re.escape(d)}(\b|[^A-Z0-9ÁÉÍÓÚÜÑ])", s) is not None

def row_day_positions(cells):
    pos = {d: [] for d in DAYS}
    for cidx, v in enumerate(cells, start=1):
        if v is None: continue
        t = str(v)
        for d in DAYS:
            if day_token_match(t, d):
                pos[d].append(cidx)
                break
    return pos

def main(xlsx, sheet_name="Segundo cuatrimestre", max_gap=30, lookback_hour=10, max_rows=3000):
    wb = load_workbook(xlsx, data_only=True)
    ws = wb[sheet_name]
    print(f"== Hoja: {ws.title} ==")
    found_groups = 0

    def looks_like_time_col(header_row, col):
        hits = 0
        rx = re.compile(r"(?i)\b([01]?\d|2[0-3])[:\.h]?[0-5]\d(?:\s*[-–—]\s*([01]?\d|2[0-3])[:\.h]?[0-5]\d)?\b")
        for r in range(header_row+1, min(ws.max_row, header_row+1+16)+1):  # 16 filas por debajo
            cell = ws.cell(r, col)
            val = cell.value
            if val is None: continue
            if getattr(cell, "is_date", False): hits += 1; continue
            if rx.search(str(val).strip()): hits += 1
        return hits >= 3

    for r in range(1, min(ws.max_row, max_rows)+1):
        row = [c.value for c in ws[r]]
        pos = row_day_positions(row)
        if not all(pos[d] for d in DAYS): 
            continue

        # prueba TODOS los LUNES como inicio de grupo
        for lun in pos["LUNES"]:
            seq = {"LUNES": lun}
            cur = lun
            ok = True
            for d in DAYS[1:]:
                cand = [c for c in pos[d] if c > cur and (c - cur) <= max_gap]
                if not cand:
                    ok = False; break
                nxt = min(cand); seq[d] = nxt; cur = nxt
            if not ok: 
                continue

            # Busca columna hora a la izquierda (1..lookback_hour)
            hcol = None
            for k in range(1, lookback_hour+1):
                c = seq["LUNES"] - k
                if c >= 1 and looks_like_time_col(r, c):
                    hcol = c; break

            print(f"[row {r}] group found cols={seq} hour_col={hcol}")
            found_groups += 1
            if found_groups >= 10:
                break
        if found_groups >= 10: break

    if found_groups == 0:
        print("NO GROUPS FOUND with these tolerances.")

if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv)>1 else r"D:\TFG\Horarios\Grado\HORARIOS 2025-26.xlsx"
    sheet = sys.argv[2] if len(sys.argv)>2 else "Segundo cuatrimestre"
    main(xlsx, sheet)
