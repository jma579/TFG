import openpyxl
import logging
import time
import re
import unicodedata
import datetime as _dt
from pathlib import Path
from typing import List, Optional, Dict, Any

from core.extraccion.horarios.entities import RawTable, CleanTable, ExtractionResult
from core.extraccion.common.entities import ExtractionMetadata, ExtractionQuality, ProcessingStatus, Warning

from core.extraccion.horarios.constants import (
    DEFAULT_EXCEL_EXTRACTOR_CONFIG, DAY_ALIASES, DAYS_CANONICAL,
    NORMALIZE_LINE_SEPS, RE_MULTI_SPACE,TIME_LIKE_REGEX,
    RE_GRUPO_PL, RE_GRUPO_PA, RE_GRUPO_GENERIC,
    RE_AULA, RE_AULA_LAB, RE_AULA_LSC, RE_AULA_SEMINARIO, RE_AULA_ABBREV,
    MODALIDAD_KEYWORDS, UNKNOWN_TOKENS,
    QUALITY_GOOD_CELL_COVERAGE, QUALITY_ACCEPTABLE_CELL_COVERAGE, QUALITY_POOR_CELL_COVERAGE,
    CONFIDENCE_CELL_COVERAGE, CONFIDENCE_NO_TEXT_PENALTY,
    QUALITY_CELL_MIN_CHARS, QUALITY_LONG_SESSION_MIN_STREAK,
    LOW_COHERENCE_WARNING_THRESHOLD, HEADER_MAX_DAY_GAP, HOUR_LOOKBACK_MAX,
    TIME_COL_VALIDATION_ROWS, TIME_COL_MIN_MATCHES
)

from core.extraccion.common.constants import (
    WEIGHT_BASIC_METRICS, WEIGHT_ACADEMIC_PATTERNS, WEIGHT_QUALITY_INDICATORS,
    BASIC_WEIGHT_STRUCTURE, BASIC_WEIGHT_CHAR_QUALITY, BASIC_WEIGHT_WORD_QUALITY,
    ACADEMIC_WEIGHT_CODES, ACADEMIC_WEIGHT_TERMINOLOGY, ACADEMIC_WEIGHT_SCHEDULE,
    QUALITY_WEIGHT_COHERENCE, QUALITY_WEIGHT_ERROR_ABSENCE,
    THRESHOLD_EXCELLENT, THRESHOLD_GOOD, THRESHOLD_ACCEPTABLE, THRESHOLD_POOR,
    BONUS_ACADEMIC_EXCELLENCE, BONUS_SOLID_STRUCTURE,
    PENALTY_HIGH_NOISE, PENALTY_CORRUPTION,
    THRESHOLD_STRUCTURE_EXCELLENCE, THRESHOLD_HIGH_NOISE_LEVEL, THRESHOLD_SIGNIFICANT_CORRUPTION,
    THRESHOLD_MULTIPLE_SUBJECT_CODES,
    MIN_CONFIDENCE,
)

class ExcelScheduleExtractor:
    def __init__(self, config: Optional[Dict[str, Any]] = None):

        # 1. Configurar logging con nivel personalizable
        self.logger = logging.getLogger(__name__)

        # 2. Aplicar configuración personalizada
        self.config = DEFAULT_EXCEL_EXTRACTOR_CONFIG.copy()
        if config:
            self.config.update(config)

        # 3. Configurar nivel de logging si se especifica
        log_level = getattr(logging, self.config.get('log_level', 'INFO').upper(), logging.INFO)
        self.logger.setLevel(log_level)
        if not self.logger.handlers:
            console_handler = logging.StreamHandler()
            console_handler.setLevel(log_level)
            formatter = logging.Formatter(
                '%(levelname)s [%(name)s] %(message)s'
            )
            console_handler.setFormatter(formatter)
            self.logger.addHandler(console_handler)
        
        # 4. Inicializar estadísticas
        self.stats = {
            # === Contadores de extracción ===
            'extractions_total': 0,           # Total de archivos procesados
            'extractions_success': 0,         # Extracciones exitosas
            'extractions_failed': 0,          # Extracciones fallidas
            
            # === Contadores de bloques ===
            'blocks_detected': 0,             # Total de bloques detectados
            'blocks_extracted': 0,            # Bloques extraídos con éxito
            'blocks_rejected': 0,             # Bloques rechazados por calidad
            
            # === Contadores de hojas ===
            'sheets_processed': 0,            # Total de hojas procesadas
            'sheets_with_schedules': 0,       # Hojas que contenían horarios
            'sheets_empty': 0,                # Hojas vacías o sin horarios
            
            # === Métricas de calidad ===
            'avg_quality_score': 0.0,         # Score promedio de calidad
            'avg_cell_coverage': 0.0,         # Cobertura promedio de celdas útiles
            'avg_blocks_per_file': 0.0,       # Promedio de bloques por archivo
            
            # === Métricas de tiempo ===
            'avg_processing_time': 0.0,       # Tiempo promedio de procesamiento
            'total_processing_time': 0.0,     # Tiempo total acumulado
            
            # === Detección de estructura ===
            'merged_cells_detected': 0,       # Total de celdas combinadas detectadas
            'sessions_with_spans': 0,         # Sesiones con duración > 60 min
            
            # === Errores y advertencias ===
            'warnings_total': 0,              # Total de advertencias generadas
            'errors_total': 0,                # Total de errores capturados
        }
        
        self.logger.info("ExcelScheduleExtractor inicializado correctamente")


    def extract(self, path: str) -> ExtractionResult:
        """
        Orquesta el proceso completo de extracción.
        """
        start_time = time.time()
        self.stats['extractions_total'] += 1

        try:
            # 1. Validaciones de entrada
            self._validate_input(path)
            self.logger.info(f"Iniciando extracción de: {path}")

            # 2. Cargar workbook y buscar bloques
            self._load_workbook(path)
            blocks = self._find_blocks()
            if not blocks:
                self.stats['extractions_failed'] += 1
                raise ValueError(
                    "No se encontraron bloques de horarios válidos en el Excel. "
                    "Verifique que el archivo contenga tablas con días de la semana y horas."
                )
            self.logger.debug(f"Detectados {len(blocks)} bloques de horarios")
            
            # 3. Extraer tablas por cada bloque
            raw_tables = []
            clean_tables = []
            for i, block in enumerate(blocks):
                try:
                    raw = self._extract_raw_table(block)
                    clean = self._extract_clean_table(raw)
                    raw_tables.append(raw)
                    clean_tables.append(clean)
                    self.stats['blocks_extracted'] += 1
                except Exception as e:
                    self.logger.warning(f"Error extrayendo bloque {i+1}: {e}")
                    self.stats['blocks_rejected'] += 1
                    continue
            if not clean_tables:
                self.stats['extractions_failed'] += 1
                raise ValueError(
                    "Todos los bloques detectados fueron rechazados por baja calidad. "
                    "Verifique la estructura de las tablas en el Excel."
                )
            
            # 4. Evaluar calidad global
            quality_score, cell_coverage = self._evaluate_extraction_quality(clean_tables)
            quality, confidence = self._map_quality_to_enum(quality_score, cell_coverage)
            
            # 5. Verificar umbral mínimo
            if quality == ExtractionQuality.UNUSABLE:
                self.stats['extractions_failed'] += 1
                raise ValueError(
                    "Excel no contiene estructura de horarios con calidad suficiente. "
                    f"Score de calidad: {quality_score:.2f}, Cobertura: {cell_coverage:.2f}. "
                    "Se requiere al menos POOR para continuar."
                )
            
            # 6. Construcción de metadatos
            processing_time = time.time() - start_time
            
            metadata = self._build_success_metadata(
                quality, confidence, clean_tables, processing_time, path
            )
            
            # 7. Actualizar estadísticas de éxito
            self.stats['extractions_success'] += 1
            self._update_stats_success(processing_time, len(blocks), quality_score, cell_coverage)
            
            # 8. Log y retorno
            self.logger.info(
                f"Extracción Excel completada: {quality.value}, "
                f"{confidence:.2f} confianza, {len(clean_tables)} tablas extraídas"
            )
            
            return ExtractionResult(
                raw_tables=raw_tables,
                clean_tables=clean_tables,
                extraccion_metadata=metadata
            )
        except Exception as e:
            return self._handle_extraction_error(e, path, start_time)


    def _validate_input(self, path: str):
        """
        Validar archivo Excel de entrada.
        
        Raises:
            FileNotFoundError: Si el archivo no existe
            ValueError: Si no es un Excel válido
        """
        file = Path(path)

        # 1. Verificar que el archivo existe
        if not file.exists():
            raise FileNotFoundError(f"Excel no encontrado: {path}")
        
        # 2. Validar extensión
        valid_extensions = {'.xlsx', '.xlsm'}
        if file.suffix.lower() not in valid_extensions:
            raise ValueError(
                f"Extensión no válida: {file.suffix}. "
                f"Se requiere Excel moderno: {', '.join(valid_extensions)}"
            )
        
        # 3. Validar que es un Excel válido (magic bytes)
        try:
            with open(path, 'rb') as f:
                header = f.read(4)
                # Excel (.xlsx, .xlsm) son archivos ZIP (magic bytes: 50 4B 03 04)
                if not header.startswith(b'PK\x03\x04'):
                    raise ValueError(
                        "El archivo no tiene formato Excel válido. "
                        "Los archivos .xlsx/.xlsm son contenedores ZIP."
                    )
        except IOError as e:
            raise ValueError(f"Error al leer archivo Excel: {e}")

        # 4. Validar que se puede abrir con openpyxl (test rápido)
        try:
            # Intentar abrir sin cargar datos (solo estructura)
            wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
            
            # Verificar que tiene al menos una hoja
            if len(wb.sheetnames) == 0:
                raise ValueError("El Excel no contiene hojas de cálculo")
            
            wb.close()
            
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError(
                "El archivo está corrupto o no es un Excel válido. "
                "Verifique que se puede abrir con Excel/LibreOffice."
            )
        except Exception as e:
            raise ValueError(f"Error validando estructura Excel: {e}")
        
        self.logger.debug(f"Validación exitosa: {file.name}")
    
    def _load_workbook(self, path: str):
        """
        Carga el archivo Excel en memoria.
        
        Args:
            path: Ruta al archivo Excel
            
        Raises:
            ValueError: Si el workbook no se puede cargar o está vacío
        """
        try:
            # Cargar con data_only=True para leer valores, no fórmulas
            self.workbook = openpyxl.load_workbook(path, data_only=True)
            self.filepath = path
            
            # Validar que tiene hojas
            if not self.workbook.sheetnames:
                raise ValueError("El Excel no contiene hojas")
            
            self.logger.info(
                f"Workbook cargado: {len(self.workbook.sheetnames)} hojas "
                f"({', '.join(self.workbook.sheetnames)})"
            )
            
        except Exception as e:
            raise ValueError(f"Error cargando workbook: {e}")
        
    def _find_blocks(self) -> List[Dict[str, Any]]:
        """
        Localiza todos los bloques de horarios en todas las hojas.
        
        Returns:
            Lista de diccionarios con información de cada bloque detectado
        """
        blocks = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            self.stats['sheets_processed'] += 1
            
            self.logger.debug(f"Escaneando hoja: {sheet_name}")
            
            # Buscar bloques en esta hoja
            sheet_blocks = self._find_blocks_in_sheet(sheet, sheet_name)
            
            if sheet_blocks:
                self.stats['sheets_with_schedules'] += 1
                blocks.extend(sheet_blocks)
                self.logger.info(
                    f"Hoja '{sheet_name}': {len(sheet_blocks)} bloques detectados"
                )
            else:
                self.stats['sheets_empty'] += 1
                self.logger.warning(f"Hoja '{sheet_name}': sin bloques detectados")
        
        self.stats['blocks_detected'] = len(blocks)
        self.logger.info(f"Total de bloques detectados: {len(blocks)}")
        
        return blocks
    
    def _extract_raw_table(self, block: dict) -> RawTable:
        """
        Construye RawTable del bloque Excel y precalcula:
        - row_hour_ranges: rango base (inicio, fin) por fila de datos
        - merge_span_matrix: altura/continuación del merge por (fila, día)
        """
        sheet = block['sheet']
        sheet_name = block['sheet_name']
        time_col = block['time_col']
        day_cols_map = block['day_cols']  # { "LUNES": col, ... }
        data_start = block['data_start_row']
        data_end = block['data_end_row']

        header = ["HORA"] + list(DAYS_CANONICAL)
        data: list[list[str]] = [header]

        # rangos base por fila (solo datos); matriz de merges (filas x 5)
        row_hour_ranges: list[Optional[tuple[str, str]]] = []
        merge_span_matrix: list[list[int]] = []

        for row in range(data_start, data_end + 1):
            # columna HORA (texto)
            hour_txt = self._cell_text(sheet, row, time_col)
            # calcula rango base
            base_range = self._extract_time_range_from_text(hour_txt)
            if not base_range:
                t0 = self._parse_time(hour_txt)
                if t0:
                    base_range = (t0, self._minutes_to_hhmm(self._time_to_minutes(t0) + 60))
                else:
                    base_range = None
            row_hour_ranges.append(base_range)

            # fila cruda (hora + 5 días)
            row_vals = [hour_txt]
            merges_row: list[int] = []

            for day in DAYS_CANONICAL:
                col = day_cols_map[day]
                row_vals.append(self._cell_text(sheet, row, col))
                merges_row.append(self._get_merged_rows_span(sheet, row, col))

            data.append(row_vals)
            merge_span_matrix.append(merges_row)

        return RawTable(
            data=data,
            source="excel",
            sheet=sheet_name,
            page=None,
            lane_index=block.get("lane_index"),
            block_id=block.get("block_id") or f"{sheet_name}!R{block['header_row']}",
            header_row=block["header_row"],
            data_start_row=data_start,
            data_end_row=data_end,
            time_col=time_col,
            day_cols=[block['day_cols'][d] for d in DAYS_CANONICAL],
            row_hour_ranges=row_hour_ranges,
            merge_span_matrix=merge_span_matrix,
            titulacion=block.get("titulacion"),
            curso=block.get("curso"),
            mencion=block.get("mencion"),
            extra={}
        )
    
    def _extract_clean_table(self, raw: RawTable) -> CleanTable:
        """
        Normaliza RawTable a rejilla 30':
        - Rango en celda del día (si existe) > rango base de fila (si existe)
        - Merge vertical por día extiende fin (múltiplos de la base)
        """
        days = list(DAYS_CANONICAL)
        time_axis = self._build_time_axis_30min()
        n_intervals = len(time_axis) - 1
        cells = [["" for _ in range(5)] for _ in range(n_intervals)]

        data = raw.data
        # Filas de datos empiezan en idx 1 (0 es cabecera)
        num_rows = len(data) - 1

        for r_idx in range(num_rows):
            row = data[r_idx + 1]
            hour_text = row[0]
            base_range = None
            if raw.row_hour_ranges and r_idx < len(raw.row_hour_ranges):
                base_range = raw.row_hour_ranges[r_idx]
            if not base_range:
                # fallback por si el raw no lo trajo (no debería ocurrir)
                base_range = self._extract_time_range_from_text(hour_text)
                if not base_range:
                    t0 = self._parse_time(hour_text)
                    if t0:
                        base_range = (t0, self._minutes_to_hhmm(self._time_to_minutes(t0) + 60))

            # Para cada día (5 columnas)
            for d_idx in range(5):
                day_text = row[1 + d_idx].strip()
                if not day_text:
                    continue

                # rango del día tiene prioridad
                day_range = self._extract_time_range_from_text(day_text)
                if day_range:
                    start, end = day_range
                    base_minutes = self._time_to_minutes(end) - self._time_to_minutes(start)
                    if base_minutes <= 0:
                        base_minutes = 60
                elif base_range:
                    start, end = base_range
                    base_minutes = self._time_to_minutes(end) - self._time_to_minutes(start)
                    if base_minutes <= 0:
                        base_minutes = 60
                else:
                    # sin tiempos, no podemos ubicar slot
                    continue

                # Extender por merge vertical según raw.merge_span_matrix
                if raw.merge_span_matrix and r_idx < len(raw.merge_span_matrix):
                    span = raw.merge_span_matrix[r_idx][d_idx]
                    if span and span > 1:
                        end = self._minutes_to_hhmm(self._time_to_minutes(end) + (span - 1) * base_minutes)

                # Pintar slots
                for i in self._find_interval_indices(time_axis, start, end):
                    cells[i][d_idx] = day_text

        # row_spans por continuidad
        row_spans = [1] * n_intervals
        for i in range(n_intervals):
            any_start = False
            any_cont = False
            for d in range(5):
                cur = (cells[i][d] or "").strip()
                if not cur:
                    continue
                prev_eq = (i > 0 and (cells[i - 1][d] or "").strip() == cur)
                next_eq = (i + 1 < n_intervals and (cells[i + 1][d] or "").strip() == cur)
                if not prev_eq and next_eq:
                    any_start = True
                elif prev_eq:
                    any_cont = True
            if any_start:
                row_spans[i] = 2
            elif any_cont:
                row_spans[i] = 0
            else:
                row_spans[i] = 1

        return CleanTable(
            time_axis=time_axis,
            days=days,
            cells=cells,
            source=raw.source,
            sheet=raw.sheet,
            page=raw.page,
            lane_index=raw.lane_index,
            block_id=raw.block_id,
            titulacion=raw.titulacion,
            curso=raw.curso,
            mencion=raw.mencion,
            row_spans=row_spans
        )

    def _evaluate_extraction_quality(self, clean_tables: list[CleanTable]) -> tuple[float, float]:
        """
        Calcula un score de calidad (0..1) con pesos de constants_common
        y devuelve también la cobertura de celdas (0..1).
        Registra en self.stats métricas auxiliares (sesiones largas, etc.).
        """
        if not clean_tables:
            return 0.0, 0.0

        # Compila regex de dominio una vez
        re_time_like = re.compile(TIME_LIKE_REGEX, re.IGNORECASE)
        re_codes = [
            re.compile(RE_GRUPO_PL, re.IGNORECASE),
            re.compile(RE_GRUPO_PA, re.IGNORECASE),
            re.compile(RE_GRUPO_GENERIC, re.IGNORECASE),
        ]
        re_aulas = [
            re.compile(RE_AULA, re.IGNORECASE),
            re.compile(RE_AULA_LAB, re.IGNORECASE),
            re.compile(RE_AULA_LSC, re.IGNORECASE),
            re.compile(RE_AULA_SEMINARIO, re.IGNORECASE),
            re.compile(RE_AULA_ABBREV, re.IGNORECASE),
        ]
        modalidad_tokens = set().union(*MODALIDAD_KEYWORDS.values())

        # Acumuladores globales
        total_slots = 0
        filled_slots = 0
        valid_tables = 0
        long_session_streaks = 0
        cells_with_codes = 0
        cells_with_aulas = 0
        cells_with_modalidad = 0
        cells_with_time_like = 0
        cells_legible = 0
        cells_wordy = 0
        cells_noise = 0

        for ct in clean_tables:
            # Estructura válida
            is_valid_days = list(ct.days) == list(DAYS_CANONICAL)
            is_valid_axis = self._is_valid_axis(ct.time_axis)
            if is_valid_days and is_valid_axis:
                valid_tables += 1

            n_intervals = max(0, len(ct.time_axis) - 1)
            total_slots += n_intervals * 5
            if n_intervals == 0:
                continue

            # Coherencia: slots con vecino vertical igual
            table_filled = 0
            table_coherent = 0

            for d in range(5):
                streak = 0
                prev = None
                for i in range(n_intervals):
                    cur = (ct.cells[i][d] or "").strip()
                    if cur:
                        table_filled += 1
                        prev_eq = (i > 0 and (ct.cells[i-1][d] or "").strip() == cur)
                        next_eq = (i+1 < n_intervals and (ct.cells[i+1][d] or "").strip() == cur)
                        if prev_eq or next_eq:
                            table_coherent += 1

                        # streaks de sesiones largas (>= QUALITY_LONG_SESSION_MIN_STREAK)
                        if prev == cur:
                            streak += 1
                        else:
                            if streak + 1 >= QUALITY_LONG_SESSION_MIN_STREAK:
                                long_session_streaks += 1
                            streak = 1
                            prev = cur

                        # calidad de caracteres/palabras y señales académicas
                        text = cur
                        if len(text) >= QUALITY_CELL_MIN_CHARS:
                            cells_legible += 1
                        if re.search(r"[A-Za-zÁÉÍÓÚÜÑ0-9]", text):
                            cells_wordy += 1
                        if any(rc.search(text) for rc in re_codes):
                            cells_with_codes += 1
                        if any(ra.search(text) for ra in re_aulas):
                            cells_with_aulas += 1
                        if re_time_like.search(text):
                            cells_with_time_like += 1
                        # modalidad: presencia de token conocido
                        if any(tok in text.upper() for tok in modalidad_tokens):
                            cells_with_modalidad += 1
                        # ruido: tokens “vacíos” conocidos
                        if text.strip() in UNKNOWN_TOKENS:
                            cells_noise += 1
                    else:
                        if streak >= QUALITY_LONG_SESSION_MIN_STREAK:
                            long_session_streaks += 1
                        streak = 0
                        prev = None
                if streak >= QUALITY_LONG_SESSION_MIN_STREAK:
                    long_session_streaks += 1

            filled_slots += table_filled
            # Guardar coherencia por tabla en stats si quieres (opcional)

        # Métricas globales
        coverage = (filled_slots / total_slots) if total_slots else 0.0
        valid_ratio = (valid_tables / len(clean_tables)) if clean_tables else 0.0
        coherence = 0.0
        if filled_slots > 0:
            # “coherence” se computó por tabla; para la mezcla final lo recalculamos rápido
            # como proporción de slots ocupados que tienen vecino (aprox via contadores parciales)
            # (si quieres más exactitud, suma table_coherent global dentro del bucle)
            pass

        # Para exactitud, recontamos coherencia global en un segundo pase (barato)
        coherent_slots = 0
        for ct in clean_tables:
            n_intervals = max(0, len(ct.time_axis) - 1)
            for d in range(5):
                for i in range(n_intervals):
                    cur = (ct.cells[i][d] or "").strip()
                    if not cur:
                        continue
                    prev_eq = (i > 0 and (ct.cells[i-1][d] or "").strip() == cur)
                    next_eq = (i+1 < n_intervals and (ct.cells[i+1][d] or "").strip() == cur)
                    if prev_eq or next_eq:
                        coherent_slots += 1
        coherence = (coherent_slots / filled_slots) if filled_slots else 0.0

        # Ratios auxiliares
        legible_ratio = (cells_legible / filled_slots) if filled_slots else 0.0
        wordy_ratio = (cells_wordy / filled_slots) if filled_slots else 0.0
        code_ratio = (cells_with_codes / filled_slots) if filled_slots else 0.0
        aula_ratio = (cells_with_aulas / filled_slots) if filled_slots else 0.0
        modalidad_ratio = (cells_with_modalidad / filled_slots) if filled_slots else 0.0
        time_like_ratio = (cells_with_time_like / filled_slots) if filled_slots else 0.0
        noise_ratio = (cells_noise / filled_slots) if filled_slots else 0.0

        # 1) Básicas (WEIGHT_BASIC_METRICS)
        basic_structure = valid_ratio
        basic_char_quality = legible_ratio
        basic_word_quality = wordy_ratio
        basic_score = (
            BASIC_WEIGHT_STRUCTURE * basic_structure +
            BASIC_WEIGHT_CHAR_QUALITY * basic_char_quality +
            BASIC_WEIGHT_WORD_QUALITY * basic_word_quality
        )

        # 2) Académicas (WEIGHT_ACADEMIC_PATTERNS)
        academic_codes = code_ratio
        academic_terms = max(aula_ratio, modalidad_ratio)  # señal fuerte si hay aula o modalidad
        academic_schedule = time_like_ratio
        academic_score = (
            ACADEMIC_WEIGHT_CODES * academic_codes +
            ACADEMIC_WEIGHT_TERMINOLOGY * academic_terms +
            ACADEMIC_WEIGHT_SCHEDULE * academic_schedule
        )

        # 3) Indicadores de calidad (WEIGHT_QUALITY_INDICATORS)
        quality_coherence = coherence
        quality_error_absence = 1.0 - noise_ratio
        quality_score_grp = (
            QUALITY_WEIGHT_COHERENCE * quality_coherence +
            QUALITY_WEIGHT_ERROR_ABSENCE * quality_error_absence
        )

        # Mezcla ponderada global
        quality_score = (
            WEIGHT_BASIC_METRICS * basic_score +
            WEIGHT_ACADEMIC_PATTERNS * academic_score +
            WEIGHT_QUALITY_INDICATORS * quality_score_grp
        )

        # Bonos y penalizaciones (constants_common)
        if basic_structure >= THRESHOLD_STRUCTURE_EXCELLENCE:
            quality_score += BONUS_SOLID_STRUCTURE
        # "Excelencia académica": varias señales de códigos/terminología
        if (cells_with_codes >= THRESHOLD_MULTIPLE_SUBJECT_CODES) and (academic_terms > 0.2):
            quality_score += BONUS_ACADEMIC_EXCELLENCE
        if noise_ratio >= THRESHOLD_HIGH_NOISE_LEVEL:
            quality_score -= PENALTY_HIGH_NOISE
        # “Corrupción” aproximada: pocas tablas válidas
        if valid_ratio <= THRESHOLD_SIGNIFICANT_CORRUPTION:
            quality_score -= PENALTY_CORRUPTION

        # Bonus suaves por cobertura (constants.py)
        if coverage >= QUALITY_GOOD_CELL_COVERAGE:
            quality_score += 0.02
        elif coverage >= QUALITY_ACCEPTABLE_CELL_COVERAGE:
            quality_score += 0.01
        elif coverage < QUALITY_POOR_CELL_COVERAGE:
            quality_score -= 0.02

        # Normalización a [0,1]
        quality_score = max(0.0, min(1.0, quality_score))

        # Guarda métricas útiles
        self.stats['long_sessions'] = int(long_session_streaks)
        self.stats['valid_table_ratio'] = float(valid_ratio)
        self.stats['coverage'] = float(coverage)
        self.stats['coherence'] = float(coherence)
        self.stats['noise_ratio'] = float(noise_ratio)

        return quality_score, coverage

    def _map_quality_to_enum(self, quality_score: float, coverage: float) -> tuple[ExtractionQuality, float]:
        """
        Mapea el score a ExtractionQuality con umbrales compartidos (constants_common)
        y calcula una confianza simple basada en score y cobertura.
        """
        if quality_score >= THRESHOLD_EXCELLENT:
            level = ExtractionQuality.EXCELLENT
        elif quality_score >= THRESHOLD_GOOD:
            level = ExtractionQuality.GOOD
        elif quality_score >= THRESHOLD_ACCEPTABLE:
            level = ExtractionQuality.ACCEPTABLE
        elif quality_score >= THRESHOLD_POOR:
            level = ExtractionQuality.POOR
        else:
            level = ExtractionQuality.UNUSABLE

        # Confianza: base mínima + aportes de score y cobertura
        conf = MIN_CONFIDENCE + 0.5 * quality_score + CONFIDENCE_CELL_COVERAGE * coverage
        if coverage == 0:
            conf -= CONFIDENCE_NO_TEXT_PENALTY
        conf = max(0.0, min(0.98, conf))

        return level, conf

    def _build_success_metadata(
        self,
        quality: ExtractionQuality,
        confidence: float,
        clean_tables: List[CleanTable],
        processing_time: float,
        path: str
    ) -> ExtractionMetadata:
        """
        Metadatos de éxito para Excel, homólogos al extractor de Fichas:
        - page_count        ← hojas procesadas
        - pages_with_text   ← hojas con horarios
        - has_embedded_text ← True si hay al menos una celda no vacía
        - char_count        ← suma de longitudes de celdas no vacías
        - word_count        ← suma de palabras de celdas no vacías
        """

        # 1) Estadísticas de texto y cobertura por tabla
        filled_cells = 0
        char_count = 0
        word_count = 0

        tables_with_acceptable_coverage = 0  # por si prefieres mantener esta métrica también

        for table in clean_tables:
            n_intervals = len(table.cells)
            n_days = len(table.days)
            table_total = n_intervals * n_days
            table_filled = 0

            for row in table.cells:
                for cell_text in row:
                    s = (cell_text or "").strip()
                    if s:
                        filled_cells += 1
                        table_filled += 1
                        char_count += len(s)
                        word_count += len(s.split())

            table_cov = (table_filled / table_total) if table_total > 0 else 0.0
            if table_cov >= QUALITY_ACCEPTABLE_CELL_COVERAGE:
                tables_with_acceptable_coverage += 1

        has_embedded_text = (filled_cells > 0)

        # 2) “Páginas” = hojas (para mantener simetría con fichas)
        page_count = int(self.stats.get("sheets_processed", 0))
        pages_with_text = int(self.stats.get("sheets_with_schedules", 0))

        # 3) Warnings (en base a métricas reales y constantes)
        warnings = []

        long_sessions = int(self.stats.get("long_sessions", self.stats.get("sessions_with_spans", 0)))
        if long_sessions == 0 and page_count > 0:
            warnings.append(Warning(
                message="No se detectaron celdas combinadas; las sesiones largas podrían no estar identificadas.",
                severity="minor"
            ))

        # Calidad baja
        if quality == ExtractionQuality.POOR:
            warnings.append(Warning(
                message=(
                    f"Calidad POOR: {confidence:.2%} de confianza. "
                    f"Cobertura global: {self.stats.get('coverage', 0.0):.1%}. "
                    "Revisar tablas extraídas."
                ),
                severity="moderate",
            ))

        # Bloques rechazados
        blocks_rejected = int(self.stats.get("blocks_rejected", 0))
        if blocks_rejected > 0:
            warnings.append(Warning(
                message=f"{blocks_rejected} bloques rechazados por calidad insuficiente o estructura inválida.",
                severity="minor",
            ))

        # Hojas sin horarios
        sheets_empty = int(self.stats.get("sheets_empty", 0))
        if sheets_empty > 0:
            warnings.append(Warning(
                message=f"{sheets_empty} hojas no contenían bloques de horarios válidos.",
                severity="minor",
            ))

        # Spans ausentes
        sessions_with_spans = int(self.stats.get("sessions_with_spans", 0))
        if sessions_with_spans == 0 and page_count > 0:
            warnings.append(Warning(
                message="No se detectaron celdas combinadas; las sesiones largas podrían no estar identificadas.",
                severity="minor",
            ))

        # Cobertura global baja (usa el umbral “aceptable” del flujo de horarios)
        coverage = float(self.stats.get("coverage", 0.0))
        if coverage < QUALITY_ACCEPTABLE_CELL_COVERAGE:
            warnings.append(Warning(
                message=f"Cobertura global baja ({coverage:.1%}). Muchos intervalos sin contenido.",
                severity="moderate",
            ))

        # Coherencia vertical baja (usa constante común)
        coherence = float(self.stats.get("coherence", 0.0))
        if coherence < LOW_COHERENCE_WARNING_THRESHOLD:
            warnings.append(Warning(
                message=f"Coherencia vertical baja ({coherence:.1%}). Posibles sesiones fragmentadas.",
                severity="moderate",
            ))

        # 4) Tamaño del fichero Excel
        try:
            file_size_mb = Path(path).stat().st_size / (1024 * 1024)
        except Exception as e:
            self.logger.warning(f"No se pudo obtener tamaño del archivo: {e}")
            file_size_mb = 0.0

        # 5) Errors/warnings auxiliares
        errors_list = list(getattr(self, "errors", []))  # si no llevas errores acumulados, quedará []

        # 6) Construcción del metadata homogéneo con fichas
        metadata = ExtractionMetadata(
            quality=quality,
            confidence=confidence,
            status=ProcessingStatus.COMPLETED,
            processing_time_seconds=float(processing_time),

            page_count=page_count,              # hojas procesadas
            pages_with_text=pages_with_text,    # hojas con horarios

            file_size_mb=float(file_size_mb),

            has_embedded_text=bool(has_embedded_text),
            char_count=int(char_count),
            word_count=int(word_count),

            errors=errors_list,
            warnings=warnings,
        )
        return metadata

    def _update_stats_success(self, processing_time: float, blocks_total: int, quality_score: float, coverage: float) -> None:
        """
        Consolida métricas acumuladas tras una extracción exitosa.
        - Actualiza medias incrementales de calidad, cobertura, tiempo y bloques por archivo.
        - Acumula sesiones largas detectadas (spans).
        - Suma advertencias/errores si se mantienen en self.warnings/self.errors.
        """
        # n de éxitos ya incrementado en extract() antes de llamarnos
        success_n = int(self.stats.get("extractions_success", 0))
        if success_n <= 0:
            # fallback defensivo (no debería ocurrir)
            success_n = 1

        # --- Medias incrementales ---
        # Calidad
        prev_q = float(self.stats.get("avg_quality_score", 0.0))
        self.stats["avg_quality_score"] = (prev_q * (success_n - 1) + float(quality_score)) / success_n

        # Cobertura
        prev_cov = float(self.stats.get("avg_cell_coverage", 0.0))
        self.stats["avg_cell_coverage"] = (prev_cov * (success_n - 1) + float(coverage)) / success_n

        # Bloques por archivo
        prev_blocks = float(self.stats.get("avg_blocks_per_file", 0.0))
        self.stats["avg_blocks_per_file"] = (prev_blocks * (success_n - 1) + float(blocks_total)) / success_n

        # Tiempos
        prev_time_avg = float(self.stats.get("avg_processing_time", 0.0))
        self.stats["avg_processing_time"] = (prev_time_avg * (success_n - 1) + float(processing_time)) / success_n
        self.stats["total_processing_time"] = float(self.stats.get("total_processing_time", 0.0)) + float(processing_time)

        # --- Agregados útiles ---
        # Pasar la métrica puntual de esta corrida a un acumulado estable
        long_sessions_this_run = int(self.stats.get("long_sessions", 0))
        self.stats["sessions_with_spans"] = int(self.stats.get("sessions_with_spans", 0)) + long_sessions_this_run
        # (opcional) limpia la métrica puntual si no la usas después
        # self.stats["long_sessions"] = 0

        # Sumar advertencias/errores si mantienes buffers en el objeto
        try:
            warnings_count = len(getattr(self, "warnings", []))
            errors_count = len(getattr(self, "errors", []))
            self.stats["warnings_total"] = int(self.stats.get("warnings_total", 0)) + int(warnings_count)
            self.stats["errors_total"] = int(self.stats.get("errors_total", 0)) + int(errors_count)
        except Exception:
            # tolerante: no bloquea si no existen esos atributos
            pass

    def _handle_extraction_error(self, error: Exception, path: str, start_time: float) -> ExtractionResult:
        """
        Manejo centralizado de errores en extracción Excel.
        Devuelve ExtractionResult con metadatos FAILED y listas vacías de tablas.
        Evita doble-contar fallos si ya se registró previamente en esta ejecución.
        """
        # Evitar doble conteo si ya se subió en alguna rama previa
        if not hasattr(self, "_failure_recorded") or not getattr(self, "_failure_recorded"):
            try:
                self.stats["extractions_failed"] = int(self.stats.get("extractions_failed", 0)) + 1
            except Exception:
                # tolerante: nunca bloquear por stats
                pass
            self._failure_recorded = True

        # Log del error
        try:
            self.logger.error(f"Error en extracción Excel: {error}")
        except Exception:
            pass

        # Calcular tamaño de archivo de forma segura
        try:
            file_size_mb = Path(path).stat().st_size / (1024 * 1024)
        except Exception:
            file_size_mb = 0.0

        # “Páginas” ↔ hojas (mapeo coherente con flujo de fichas)
        page_count = int(self.stats.get("sheets_processed", 0))
        pages_with_text = 0

        # Construcción de warnings y errors
        try:
            from core.extraccion.common.entities import Warning
        except Exception:
            Warning = None  # fallback: será lista de strings

        warnings_list = []
        if Warning:
            warnings_list.append(Warning(message=f"Error de extracción: {str(error)}", severity="severe"))
        else:
            warnings_list.append(f"Error de extracción: {str(error)}")

        errors_list = [str(error)]

        # Metadatos FAILED (homólogo a fichas, con semántica Excel)
        metadata = ExtractionMetadata(
            quality=ExtractionQuality.UNUSABLE,
            confidence=0.0,
            status=ProcessingStatus.FAILED,
            processing_time_seconds=(time.time() - start_time),
            page_count=page_count,           # hojas vistas hasta el fallo
            pages_with_text=pages_with_text, # 0 en error
            file_size_mb=float(file_size_mb),
            has_embedded_text=False,
            char_count=0,
            word_count=0,
            errors=errors_list,
            warnings=warnings_list,
        )

        # Sumar contadores globales de errores (tolerante si no existen buffers)
        try:
            self.stats["errors_total"] = int(self.stats.get("errors_total", 0)) + 1
        except Exception:
            pass

        # Resultado de extractor con payload vacío y metadata de error
        return ExtractionResult(
            raw_tables=[],
            clean_tables=[],
            extraccion_metadata=metadata
        )


    def _find_blocks_in_sheet(self, sheet, sheet_name: str) -> list[dict]:
        blocks = []
        max_row = sheet.max_row
        row_idx = 1
        processed_headers = set()  # ← Evita procesar la misma cabecera dos veces
        
        while row_idx <= max_row:
            if row_idx in processed_headers:
                row_idx += 1
                continue
                
            row_cells = [cell.value for cell in sheet[row_idx]]
            groups = self._find_day_groups_in_row(row_cells)
            
            if not groups:
                row_idx += 1
                continue
            
            # Procesar TODOS los grupos encontrados en esta fila
            extracted_blocks = []
            for days_found in groups:
                if len(days_found) < 5:
                    continue
                block = self._extract_block_info(sheet, sheet_name, row_idx, days_found, max_row)
                if block:
                    blocks.append(block)
                    extracted_blocks.append(block)
            
            # Marcar esta fila como procesada
            processed_headers.add(row_idx)
            
            # Avanzar SOLO si no se extrajo ningún bloque válido
            if not extracted_blocks:
                row_idx += 1
            else:
                # Saltar al final del bloque MÁS LARGO extraído
                # (para buscar la SIGUIENTE cabecera de bloque)
                max_end = max(b["data_end_row"] for b in extracted_blocks)
                row_idx = max_end + 1
        
        return blocks

    def _extract_block_info(self, sheet, sheet_name: str, header_row: int, days_found: Dict[str, int], max_row: int) -> Optional[Dict[str, Any]]:
        """
        Extrae información completa de un bloque.
        Requiere 5 días canónicos en orden y contiguos.
        Valida que la columna inmediatamente a la izquierda es de 'hora'.
        """
        self.logger.debug(f"[{sheet_name}!R{header_row}] Procesando grupo con días: {days_found}")

        # 1) Orden canon L->V y exigir 5 días
        ordered = []
        for d in DAYS_CANONICAL:
            col = days_found.get(d)
            if not col:
                self.logger.debug(f"[{sheet_name}!R{header_row}] ❌ RECHAZADO: Falta día '{d}' en cabecera")
                return None
            ordered.append((d, col))

        # 2) Contigüidad estricta de columnas (L..V consecutivos)
        ordered_cols = [c for _, c in ordered]
        diffs = [ordered_cols[i+1] - ordered_cols[i] for i in range(len(ordered_cols)-1)]
        if any(d <= 0 or d > HEADER_MAX_DAY_GAP for d in diffs):
            self.logger.debug(
                f"[{sheet_name}!R{header_row}] ❌ RECHAZADO: Gaps inválidos entre días. "
                f"Columnas={ordered_cols}, Diffs={diffs}, Max permitido={HEADER_MAX_DAY_GAP}"
            )
            return None

        lunes_col = ordered_cols[0]
        # busquemos la columna de hora a la izquierda con lookback
        candidate_hour_cols = [lunes_col - k for k in range(1, HOUR_LOOKBACK_MAX + 1) if (lunes_col - k) >= 1]
        self.logger.debug(f"[{sheet_name}!R{header_row}] Buscando columna de hora. Candidatos: {candidate_hour_cols}")
        
        time_col = None
        for hc in candidate_hour_cols:
            if self._col_looks_like_time(sheet, hc, header_row + 1):
                time_col = hc
                self.logger.debug(f"[{sheet_name}!R{header_row}] ✅ Columna de hora detectada: col {hc}")
                break
        
        if time_col is None:
            self.logger.debug(
                f"[{sheet_name}!R{header_row}] ❌ RECHAZADO: No se encontró columna de hora válida. "
                f"Probados: {candidate_hour_cols}"
            )
            return None

        # 3) Validar ancho mínimo del bloque
        min_col = time_col
        max_col = ordered_cols[-1]
        min_cols = self.config.get('min_cols_for_block', 6)
        if (max_col - min_col + 1) < min_cols:
            self.logger.debug(
                f"[{sheet_name}!R{header_row}] ❌ RECHAZADO: Bloque demasiado estrecho. "
                f"Ancho={max_col - min_col + 1}, Mínimo requerido={min_cols}"
            )
            return None

        # 4) Validar que la columna 'hora' realmente lo parece
        if not self._col_looks_like_time(sheet, time_col, header_row + 1):
            self.logger.debug(f"[{sheet_name}!R{header_row}] ❌ RECHAZADO: Columna {time_col} no parece de hora")
            return None

        # 5) Rango vertical del bloque
        data_start_row = header_row + 1
        # Saltar primera fila si está vacía (común en horarios con celdas combinadas)
        first_row_cells = [sheet.cell(data_start_row, c).value for c in range(min_col, max_col + 1)]
        first_row_empty = all(
            cell is None or str(cell).strip() in ('', '-', '—')
            for cell in first_row_cells
        )
        if first_row_empty:
            data_start_row += 1
            self.logger.debug(f"[{sheet_name}!R{header_row}] Primera fila vacía, iniciando datos en fila {data_start_row}")
        
        data_end_row = self._find_block_end(sheet, data_start_row, max_row, min_col, max_col)
    

        # 6) Mínimo de filas
        min_rows = self.config['min_rows_for_block']
        num_rows = data_end_row - data_start_row + 1
        if num_rows < min_rows:
            self.logger.debug(
                f"[{sheet_name}!R{header_row}] ❌ RECHAZADO: Pocas filas de datos. "
                f"Encontradas={num_rows}, Mínimo requerido={min_rows}"
            )
            return None

        # 7) Construir map day_cols en orden canónico
        day_cols_map = {d: c for d, c in ordered}
        
        self.logger.info(
            f"[{sheet_name}!R{header_row}] ✅ BLOQUE VÁLIDO: "
            f"Cols {min_col}-{max_col}, Filas {data_start_row}-{data_end_row}, "
            f"Time_col={time_col}"
        )

        titulacion_info = self._extract_titulacion_from_header(sheet, header_row)
    
        return {
            'sheet_name': sheet_name,
            'sheet': sheet,
            'header_row': header_row,
            'data_start_row': data_start_row,
            'data_end_row': data_end_row,
            'time_col': time_col,
            'day_cols': day_cols_map,
            'min_col': min_col,
            'max_col': max_col,
            'titulacion': titulacion_info.get('titulacion'), 
            'curso': titulacion_info.get('curso'),           
            'mencion': titulacion_info.get('mencion'),      
        }

    def _extract_titulacion_from_header(self, sheet, header_row: int) -> dict:
        """
        Busca en las 3 filas anteriores a la cabecera:
        - Titulación: "GRADO EN MATEMÁTICAS"
        - Curso: "PRIMER CURSO", "SEGUNDO CURSO", ...
        - Mención: "MENCIÓN EN ...", opcional
        """
        titulacion = None
        curso = None
        mencion = None
        
        for row in range(max(1, header_row - 3), header_row):
            merged_text = " ".join(
                str(cell.value or "") 
                for cell in sheet[row] 
                if cell.value
            ).strip().upper()
            
            # Detectar grado/titulación
            if "GRADO EN" in merged_text:
                titulacion = merged_text
            
            # Detectar curso (PRIMER, SEGUNDO, TERCER, CUARTO)
            import re
            curso_match = re.search(r"(PRIMER|SEGUNDO|TERCER|CUARTO)\s+CURSO", merged_text)
            if curso_match:
                curso = curso_match.group(0)
            
            # Detectar mención
            mencion_match = re.search(r"MENCI[OÓ]N\s+EN\s+(.+?)(?:\s|$)", merged_text)
            if mencion_match:
                mencion = f"MENCIÓN EN {mencion_match.group(1).strip()}"
        
        return {
            'titulacion': titulacion,
            'curso': curso,
            'mencion': mencion
        }

    def _find_block_end(self, sheet, start_row: int, max_row: int, min_col: int, max_col: int) -> int:
        """
        Encuentra la última fila del bloque.
        """
        max_empty_rows = self.config['max_empty_rows_between_blocks']
        empty_row_count = 0
        last_valid_row = start_row
        
        for row_idx in range(start_row, max_row + 1):
            row_cells = [
                sheet.cell(row_idx, col).value 
                for col in range(min_col, max_col + 1)
            ]
            
            # Verificar si es una nueva cabecera (solo en rango del bloque)
            days_in_row = self._row_days_exact(row_cells)
            if len(days_in_row) < self.config['min_days_in_header']:
                days_in_row = self._row_days_loose(row_cells)
            
            if len(days_in_row) >= self.config['min_days_in_header']:
                self.logger.debug(
                    f"Nueva cabecera detectada en fila {row_idx}, "
                    f"finalizando bloque en fila {last_valid_row}"
                )
                return last_valid_row
            
            # ✅ MEJORADO: Considerar fila válida si tiene contenido significativo
            has_content = any(
                cell is not None 
                and str(cell).strip() not in ('', '-', '—')
                and len(str(cell).strip()) > 0  # ← Asegurar que no es solo espacios
                for cell in row_cells
            )
            
            # ✅ NUEVO: También verificar si la celda tiene formato (color de fondo)
            # Esto ayuda con celdas combinadas que parecen vacías
            has_formatting = False
            try:
                for col in range(min_col, max_col + 1):
                    cell = sheet.cell(row_idx, col)
                    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb != 'FFFFFFFF':
                        has_formatting = True
                        break
            except:
                pass
            
            if has_content or has_formatting:
                empty_row_count = 0
                last_valid_row = row_idx
            else:
                empty_row_count += 1
                if empty_row_count >= max_empty_rows:
                    self.logger.debug(
                        f"{max_empty_rows} filas vacías consecutivas, "
                        f"finalizando bloque en fila {last_valid_row}"
                    )
                    return last_valid_row
        
        self.logger.debug(f"Fin de hoja alcanzado, bloque termina en fila {last_valid_row}")
        return last_valid_row

    def _extract_time_range_from_text(self, s: str | None) -> tuple[str, str] | None:
        if not s:
            return None
        import re
        s = str(s)
        m = re.search(
            r"([01]?\d|2[0-3])[:\.hH]([0-5]\d)\s*[-–—]\s*([01]?\d|2[0-3])[:\.hH]([0-5]\d)",
            s
        )
        if not m:
            return None
        a = f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
        b = f"{int(m.group(3)):02d}:{int(m.group(4)):02d}"
        return (a, b)

    def _find_interval_indices(self, time_axis: list[str], start: str, end: str) -> list[int]:
        idx = []
        s = self._time_to_minutes(start)
        e = self._time_to_minutes(end)
        if e <= s:
            return idx
        for i in range(len(time_axis) - 1):
            a = self._time_to_minutes(time_axis[i])
            b = self._time_to_minutes(time_axis[i + 1])
            if a >= s and b <= e:
                idx.append(i)
        return idx

    def _get_merged_rows_span(self, sheet, row: int, col: int) -> int:
        try:
            for r in sheet.merged_cells.ranges:
                if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col:
                    if row == r.min_row and col == r.min_col:
                        return (r.max_row - r.min_row + 1)
                    return 0
        except Exception:
            pass
        return 1

    def _cell_text(self, sheet, row: int, col: int) -> str:
        val = sheet.cell(row, col).value
        return self._normalize_spaces(str(val or ""))

    #========================= Métodos auxiliares (placeholders) ========================#
    def _norm(self, s: str) -> str:
        """Upper + strip + colapso espacios + sin acentos básicos."""
        s = (s or "").strip()
        s = " ".join(s.split())
        s = s.upper()
        s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
        return s

    def _row_days_exact(self, row_cells: list[str]) -> dict:
        """
        Devuelve {canonical_day: col_idx} si la celda es EXACTAMENTE un día (por alias).
        Conserva la PRIMERA aparición de cada día en la fila.
        """
        found = {}
        for col_idx, v in enumerate(row_cells, start=1):
            if v is None:
                continue
            cell = self._norm(str(v))
            for alias, canonical in DAY_ALIASES.items():
                if self._norm(alias) == cell:
                    if canonical not in found:          # <<< evita pisar la primera
                        found[canonical] = col_idx
                    break
        return found
    
    def _row_day_positions(self, row_cells: list[str]) -> dict[str, list[int]]:
        """ Devuelve TODAS las columnas donde aparece cada día en la fila (coincidencia laxa). """
        positions = {d: [] for d in DAYS_CANONICAL}
        for col_idx, v in enumerate(row_cells, start=1):
            if v is None:
                continue
            txt = str(v)
            for d in DAYS_CANONICAL:
                if self._day_token_match(txt, d):
                    positions[d].append(col_idx)
                    break
        return positions

    def _find_day_groups_in_row(self, row_cells: list[str]) -> list[dict[str, int]]:
        """
        Encuentra todas las tandas L-M-X-J-V NO SOLAPADAS en la fila.
        """
        pos = self._row_day_positions(row_cells)
        groups = []
        used_cols = set()  # Evitar reutilizar columnas
        
        for col_lunes in sorted(pos["LUNES"]):
            if col_lunes in used_cols:  # ← Ya usado en otro grupo
                continue
                
            seq = {"LUNES": col_lunes}
            cur = col_lunes
            ok = True
            
            for d in DAYS_CANONICAL[1:]:  # MARTES...VIERNES
                candidates = [
                    c for c in pos[d] 
                    if c > cur 
                    and (c - cur) <= HEADER_MAX_DAY_GAP
                    and c not in used_cols  # No reutilizar
                ]
                if not candidates:
                    ok = False
                    break
                nxt = min(candidates)
                seq[d] = nxt
                cur = nxt
            
            if ok:
                groups.append(seq)
                # Marcar columnas usadas
                for col in seq.values():
                    used_cols.add(col)
        
        return groups

    def _day_token_match(self, cell_text: str, canonical_day: str) -> bool:
        """
        Coincidencia laxa: la celda empieza por el día canónico (LUNES, MARTES, …)
        y después hay fin de palabra o un separador no alfanumérico.
        Soporta sufijos como '(G1)', notas, saltos de línea normalizados, etc.
        """
        s = self._norm(cell_text or "")
        d = self._norm(canonical_day or "")
        if not s or not d:
            return False
        if s == d:
            return True
        import re
        return re.match(rf"^{re.escape(d)}(\b|[^A-Z0-9ÁÉÍÓÚÜÑ])", s) is not None

    def _row_days_loose(self, row_cells: list[str]) -> dict[str, int]:
        """
        Devuelve {día_canónico: primera_col} usando coincidencia laxa (_day_token_match).
        Conserva la PRIMERA aparición de cada día en la fila.
        """
        found = {}
        for col_idx, v in enumerate(row_cells, start=1):
            if v is None:
                continue
            txt = str(v)
            for d in DAYS_CANONICAL:
                if d not in found and self._day_token_match(txt, d):
                    found[d] = col_idx
                    break
        return found
    
    def _col_looks_like_time(self, sheet, col: int, start_row: int,
                         rows: int | None = None, min_hits: int | None = None) -> bool:
        """
        ¿La columna parece de hora? Busca 'min_hits' coincidencias en 'rows' filas
        por debajo del header. Acepta '10:00', '10.30', '10:00-11:00', fechas Excel, etc.
        """

        rows = rows or TIME_COL_VALIDATION_ROWS
        min_hits = min_hits or TIME_COL_MIN_MATCHES

        rx = re.compile(r"(?i)\b([01]?\d|2[0-3])[:\.h]?[0-5]\d(?:\s*[-–—]\s*([01]?\d|2[0-3])[:\.h]?[0-5]\d)?\b")
        hits = 0
        r = start_row
        checked = 0
        while r <= sheet.max_row and checked < rows:
            cell = sheet.cell(r, col)
            val = cell.value
            # 1) datetime/time nativo
            if isinstance(val, (_dt.time, _dt.datetime)):
                hits += 1
            # 2) número con formato de fecha/hora
            elif getattr(cell, "is_date", False):
                hits += 1
            # 3) texto con patrón de hora o rango
            elif isinstance(val, str) and val.strip() and rx.search(val.strip()):
                hits += 1
            if hits >= min_hits:
                return True
            r += 1
            checked += 1
        return False

    def _normalize_spaces(self, text: str) -> str:
        """
        Normaliza texto para celdas:
        - Convierte None en "".
        - Sustituye saltos de línea por espacios.
        - Colapsa espacios múltiples.
        - Hace strip del resultado.
        """
        if text is None:
            return ""
        s = str(text)
        # Quita BOM u otros invisibles raros si hiciera falta
        s = s.replace("\ufeff", "")
        # Sustituye saltos de línea por espacio
        for sep in NORMALIZE_LINE_SEPS:
            s = s.replace(sep, " ")
        # Colapsa espacios múltiples
        s = re.sub(RE_MULTI_SPACE, " ", s)
        return s.strip()
    
    def _build_time_axis_30min(self) -> list[str]:
        # Ventana fija 08:00 → 20:30 (nodos cada 30’)
        axis = []
        h, m = 8, 0
        while (h, m) <= (20, 30):
            axis.append(f"{h:02d}:{m:02d}")
            m += 30
            if m >= 60:
                m -= 60
                h += 1
        return axis

    def _time_to_minutes(self, hhmm: str) -> int:
        hh, mm = hhmm.split(":")
        return int(hh) * 60 + int(mm)

    def _minutes_to_hhmm(self, mins: int) -> str:
        mins = max(0, mins)
        return f"{mins // 60:02d}:{mins % 60:02d}"

    def _parse_time(self, s: str | None) -> str | None:
        if not s:
            return None
        import re
        s = str(s).strip()
        m = re.search(r"\b([01]?\d|2[0-3])[:\.hH]([0-5]\d)\b", s)
        if not m:
            return None
        return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
    
    def _is_valid_axis(self, axis: list[str]) -> bool:
        if not axis or axis[0] != "08:00" or axis[-1] != "20:30":
            return False
        for i in range(len(axis) - 1):
            if self._time_to_minutes(axis[i+1]) - self._time_to_minutes(axis[i]) != 30:
                return False
        return True

    def _regex_any(self, patterns: list[str], text: str) -> bool:
        import re
        for p in patterns:
            if re.search(p, text, flags=re.IGNORECASE):
                return True
        return False
